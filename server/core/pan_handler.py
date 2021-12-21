import struct
import time
import json
import os

from openpyxl import load_workbook

from config import settings


class PanHandler(object):
    def __init__(self,conn):
        self.conn = conn
        self.username = None

    def login(self,*args):
        username,password = args
        workbook = load_workbook(settings.DB_FILE_PATH)
        sheet = workbook.active
        success = False
        for user in sheet.iter_rows(2):
            if username == user[0].value and password == user[1].value:
                success = True
                break
        if success:
            content = {'status':True,'data':'登陆成功'}
            self.username = username
            data = json.dumps(content)
            self.conn.send(data.encode('utf8'))
        else:
            content = {'status': False, 'data': '登陆失败，账号或者密码不一致'}
            data = json.dumps(content)
            self.conn.send(data.encode('utf8'))

    def register(self,*args):
        username,password = args
        workbook = load_workbook(settings.DB_FILE_PATH)
        sheet = workbook.active
        #检测用户是否存在
        exists = False
        for cell in sheet.iter_rows(2):
            if username == cell[0].value:
                exists = True
                break
        if exists:
            content = {'status':False,'data':'用户已存在'}
            content = json.dumps(content)
            self.conn.send(content.encode('utf8'))
            return
        # 如果不存在 就存入excel表格
        now_time = time.strftime('%Y-%m-%d %X')
        data = [username,password,now_time]
        sheet.append(data)
        workbook.save(settings.DB_FILE_PATH)

        # 创建用户的目录
        user_folder = os.path.join(settings.USER_FOLDER_DIR,username)
        os.makedirs(user_folder)
        # 回复消息
        content = {'status': True, 'data': '注册成功'}
        content = json.dumps(content)
        self.conn.send(content.encode('utf8'))

    def ls(self,folder_path = None):
        home_path = os.path.join(settings.USER_FOLDER_DIR,self.username)
        #判断用户是否输入了具体查看某个文件夹
        if not folder_path:
            file_list = '\n'.join(os.listdir(home_path))
            content = {'status':True,'data':file_list}
            data = json.dumps(content)
            self.conn.send(data.encode('utf8'))
            return
        # 如果用户输入了想要查看哪个文件夹
        target_folder = os.path.join(home_path,folder_path)
        if not os.path.exists(target_folder):
            content = {'status': False, 'data': '系统找不到指定路径'}
            data = json.dumps(content)
            self.conn.send(data.encode('utf8'))
            return
        file_list = '\n'.join(os.listdir(target_folder))
        content = {'status': True, 'data': file_list}
        data = json.dumps(content)
        self.conn.send(data.encode('utf8'))

    def upload(self,file_path):
        home_path = os.path.join(settings.USER_FOLDER_DIR,self.username)
        target_file_path = os.path.join(home_path,file_path)
        folder = os.path.dirname(target_file_path)
        if not os.path.exists(folder):
            os.makedirs(folder)
        #先接收报头(报头固定大小4) 进行反解
        header = self.conn.recv(4)
        total_size = struct.unpack('i',header)[0]

        #收到文件后存入本地
        with open(target_file_path,'wb') as f:
            recv_size = 0
            while recv_size < total_size:
                line = self.conn.recv(1024)
                f.write(line)
                recv_size += len(line)

    def download(self,remote_file_path):
        home_path = os.path.join(settings.USER_FOLDER_DIR,self.username)
        target_file_path = os.path.join(home_path,remote_file_path)
        if not os.path.exists(target_file_path):
            content = {'status':False,'data':'找不到指定路径'}
            data = json.dumps(content)
            self.conn.send(data.encode('utf8'))
            return

        total_size = os.path.getsize(target_file_path)
        header = struct.pack('i', total_size)
        self.conn.send(header)
        time.sleep(0.5)
        #发送数据
        with open(target_file_path,'rb')as f:
            for line in f:
                self.conn.send(line)
        content = {'status': True, 'data': '下载完毕'}
        data = json.dumps(content)
        self.conn.send(data.encode('utf8'))

    def execute(self):
        '''
        每次客户端发消息，都会触发到这个方法的执行
        返回False 代表关闭连接
        返回True 代表继续处理客户端的请求
        '''
        #1、接收客户端发来的消息
        cmd = self.conn.recv(1024).decode('utf8')
        if  cmd.upper() == 'Q':
            print('client quit out')
            return False

        cmd, *args = cmd.split(' ')
        method_map = {
            'login': self.login,
            'register': self.register,
            'ls': self.ls,
            'upload': self.upload,
            'download': self.download

        }
        method = method_map[cmd]
        method(*args)
        #2、执行相关的方法
        return  True
